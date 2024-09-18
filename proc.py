import pandas as pd
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
import os

# dps = FORMATO DE LOGISTICA INVERSA - DELL - UNISYS
# p/n = Unnamed: 3

def find_start_row_of_table(file, sheet_name):
    wb = load_workbook(filename=file, data_only=True)
    ws = wb[sheet_name]
    for table in ws.tables.values():
        start_row = ws[table.ref.split(':')[0]].row
        return start_row
    raise ValueError("No se encontró ninguna tabla en la hoja especificada.")

def comparar_excel(file1, file2, column_name1, start_row, column_name2=None, highlight_columns=None, word_column=None, palabra="NRP"):
    # Cargar los archivos Excel con pandas
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)
    
    # Cargar el archivo Excel original con openpyxl
    wb = load_workbook(file1)
    sheet_name = wb.active.title
    ws = wb[sheet_name]
    
    # Encontrar la fila de inicio de la tabla
    start_row = find_start_row_of_table(file1, sheet_name)

    if column_name2:
        # Comparar las columnas como pares a partir de la fila de inicio en ambos DataFrames
        pares_df1 = df1[[column_name1, column_name2]].iloc[start_row-1:].apply(tuple, axis=1)
        pares_df2 = df2[[column_name1, column_name2]].iloc[start_row-1:].apply(tuple, axis=1)
        coincidencias = pares_df1.isin(pares_df2)
    else:
        # Comparar solo la primera columna
        coincidencias = df1[column_name1].iloc[start_row-1:].isin(df2[column_name1].iloc[start_row-1:])

    # Detallar las coincidencias
    for df_index, coincide in enumerate(coincidencias, start=start_row):
        if coincide:
            print(f"Fila {df_index+1} coincide en ambas columnas.")
            value1 = df1.iloc[df_index - start_row + (start_row-1)][column_name1]
            if column_name2:
                value2 = df1.iloc[df_index - start_row + (start_row-1)][column_name2]
                print(f" - Columna DPS: {value1}")
                print(f" - Columna P/N: {value2}")
            else:
                print(f" - Columna: {value1}")

    # Resaltar las filas que coinciden en amarillo
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Convertir las columnas a resaltar en índices numéricos
    if highlight_columns:
        highlight_columns = [ord(col.upper()) - ord('A') + 1 for col in highlight_columns.split(',')]
    
    # Convertir la columna para insertar la palabra en índice numérico
    word_column_index = 5

#    for df_index, coincide in enumerate(coincidencias, start=start_row):
#        excel_index = df_index + 1  # Ajustar el índice para Excel
#        for col in range(1, ws.max_column + 1):
#            cell = ws.cell(row=excel_index, column=col)
#            if coincide:
#                if not highlight_columns or col in highlight_columns:
#                    cell.fill = fill
#            else:
#                if word_column_index and col == word_column_index:
#                    cell.value = palabra

    # Obtener el nombre del archivo original sin la extensión
    original_filename = os.path.splitext(file1.filename)[0]
    
    # Guardar el archivo en un objeto BytesIO para devolverlo sin guardarlo en disco
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, original_filename

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file1 = request.files["file1"]
        file2 = request.files["file2"]
        column_name1 = request.form["column_name1"]
        start_row = int(request.form.get("start_row", 1))
        compare_two_columns = "compare_two_columns" in request.form
        highlight_columns = request.form.get("highlight_columns")
        word_column = request.form.get("word_column")
        
        column_name2 = request.form["column_name2"] if compare_two_columns else None
        
        excel_modificado, original_filename = comparar_excel(file1, file2, column_name1, start_row, column_name2, highlight_columns, word_column)
        
        prefix = ""
        sufix = ""
        
        for item in original_filename:
            prefix += item
            if item == " ":
                break
        
        indice_guion = original_filename.find('-')
        # Verifica si el guion fue encontrado
        if indice_guion != -1:
            # Extrae la parte de la oración que sigue al guion
            sufix = original_filename[indice_guion + 1:].strip()
        else:
            indice_guion = original_filename.find('-')

        
        return send_file(excel_modificado, download_name=prefix +" SELECCION DE PARTES DEPURACION UPS - " + sufix + ".xlsx", as_attachment=True)
    
    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=False)
